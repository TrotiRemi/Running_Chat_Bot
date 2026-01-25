from pathlib import Path
from typing import List, Any

from openpyxl import load_workbook


def read_sheet_rows(xlsx_path: Path, sheet_name: str) -> List[List[Any]]:
    wb = load_workbook(xlsx_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in {xlsx_path.name}")
    ws = wb[sheet_name]
    rows = []
    for row in ws.iter_rows(values_only=True):
        if row is None:
            continue
        rows.append(list(row))
    return rows
